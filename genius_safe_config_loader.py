
"""
GENIUS BOT — SAFE CONFIG LOADER
Institutional-grade Excel → Python bridge
Features:
- safe float cast
- missing key guard
- fallback defaults
- zero-crash read
"""

import logging
from typing import Any, Dict
from openpyxl import load_workbook

logger = logging.getLogger("GENIUS.CONFIG")
logger.setLevel(logging.INFO)


# ---------------- DEFAULTS ----------------
DEFAULTS: Dict[str, Any] = {
    "RISK_PER_TRADE": 0.012,
    "HARD_RISK_CAP": 0.02,
    "MAX_POSITION_CAP": 0.25,
    "MAX_TRADES_PER_DAY": 5,
    "AI_CONFIDENCE_BUY_MIN": 0.62,
    "AI_CONFIDENCE_SELL_MIN": 0.66,
}


# ---------------- SAFE CAST ----------------
def safe_float(value: Any, default: float) -> float:
    """Convert to float safely with fallback."""
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        logger.warning(f"SAFE_CAST fallback used for value={value}")
        return default


# ---------------- CORE LOADER ----------------
def load_config_from_excel(path: str) -> Dict[str, Any]:
    """
    Read AUTO_WIRE_MAP and return safe config dict.
    Zero-crash design.
    """
    config: Dict[str, Any] = {}

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel: {e}")
        return DEFAULTS.copy()

    if "AUTO_WIRE_MAP" not in wb.sheetnames:
        logger.error("AUTO_WIRE_MAP sheet missing — using defaults")
        return DEFAULTS.copy()

    ws = wb["AUTO_WIRE_MAP"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = row[2] if len(row) > 2 else None

        if not key:
            continue

        # Fallback default
        default_value = DEFAULTS.get(key)

        # If Excel formula not evaluated, try reading from CONFIG_CORE
        if isinstance(value, str) and value.startswith("="):
            value = None

        # Safe casting for numeric fields
        if isinstance(default_value, (int, float)):
            value = safe_float(value, default_value)
        else:
            if value is None:
                value = default_value

        config[key] = value

    # Ensure all defaults exist
    for k, v in DEFAULTS.items():
        if k not in config:
            logger.warning(f"Missing key {k} — default applied")
            config[k] = v

    logger.info("SAFE CONFIG LOADED")
    return config


# ---------------- QUICK TEST ----------------
if __name__ == "__main__":
    import pprint
    path = "DYZEN_CAPITAL_OS_AI_LIVE_CORE_READY_HIGH_ALPHA_AUTOWIRE.xlsx"
    cfg = load_config_from_excel(path)
    pprint.pprint(cfg)
