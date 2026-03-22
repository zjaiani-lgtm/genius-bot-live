"""
backtest_v3.py — Genius Bot Vectorbt-Style Framework
═══════════════════════════════════════════════════════════════════════
Production-grade backtesting system. pandas/numpy-based (vectorbt-style).
vectorbt install-ის შემდეგ: pip install vectorbt  → drop-in replacement.

FEATURES:
  ┌─ 4 სტრატეგია ──────────────────────────────────────────────────┐
  │  A. Fixed TP/SL (ძველი ბოტი)                                   │
  │  B. ATR-based Regime (ახლანდელი ბოტი)                          │
  │  C. RSI + MACD filter (ახალი სიგნალები)                        │
  │  D. RSI + MACD + MTF + Dynamic Sizing (სრული ახალი სისტემა)   │
  └─────────────────────────────────────────────────────────────────┘
  ┌─ Analysis ──────────────────────────────────────────────────────┐
  │  • Walk-Forward Testing (train/test split)                      │
  │  • Parameter Optimization (TP/SL/RSI grid search)              │
  │  • Monte Carlo simulation (1000 runs)                           │
  │  • Per-symbol + Per-regime breakdown                            │
  │  • HTML + Excel report                                          │
  └─────────────────────────────────────────────────────────────────┘

გამოყენება:
  # OHLCV mode — Binance-დან ჩამოტვირთვა (საუკეთესო):
  python backtest_v3.py --mode ohlcv --symbols BTC/USDT ETH/USDT BNB/USDT

  # Order History mode — Binance ექსპორტიდან:
  python backtest_v3.py --mode history --file Binance_Order_History.xlsx

  # Optimization mode:
  python backtest_v3.py --mode optimize --symbols BTC/USDT ETH/USDT

  # Walk-Forward mode:
  python backtest_v3.py --mode walkforward --symbols BTC/USDT

შედეგი:
  backtest_v3_results.xlsx  — სრული trade-by-trade + metrics
  backtest_v3_report.html   — ვიზუალური dashboard
"""

from __future__ import annotations

import os
import sys
import math
import json
import time
import random
import logging
import argparse
import itertools
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple, Any

import numpy as np
import pandas as pd

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LOGGING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("bt3")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ENV CONFIG — ყველა ENV-ით კონტროლირებადია
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _ef(name: str, default: float) -> float:
    v = os.getenv(name)
    try:
        return float(v) if v is not None else default
    except Exception:
        return default


def _ei(name: str, default: int) -> int:
    v = os.getenv(name)
    try:
        return int(v) if v is not None else default
    except Exception:
        return default


# Costs
FEE_RT = _ef("ESTIMATED_ROUNDTRIP_FEE_PCT", 0.14)
SLIP   = _ef("ESTIMATED_SLIPPAGE_PCT",       0.05)
COST   = (FEE_RT + SLIP) / 100.0

# Strategy A — fixed
A_TP_PCT = _ef("TP_PCT", 3.0)
A_SL_PCT = _ef("SL_PCT", 1.0)

# Strategy B — ATR regime
B_ATR_TP_BULL      = _ef("ATR_MULT_TP_BULL",    3.0)
B_ATR_SL_BULL      = _ef("ATR_MULT_SL_BULL",    1.0)
B_ATR_TP_UNCERTAIN = 2.5
B_ATR_SL_UNCERTAIN = 1.0
B_MIN_TP           = _ef("MIN_NET_PROFIT_PCT",   0.20)
B_MIN_SL           = 0.20
B_MAX_TP           = 4.0
B_MAX_SL           = 1.5

# Regime thresholds
BULL_TREND_MIN    = _ef("REGIME_BULL_TREND_MIN",    0.35)
SIDEWAYS_ATR_MAX  = _ef("REGIME_SIDEWAYS_ATR_MAX",  0.18)
VOLATILE_ATR_MIN  = 1.50
BEAR_TREND_MAX    = -0.10

# RSI
RSI_PERIOD   = _ei("RSI_PERIOD", 14)
RSI_MIN      = _ef("RSI_MIN",    35.0)
RSI_MAX      = _ef("RSI_MAX",    70.0)
RSI_SELL_MIN = _ef("RSI_SELL_MIN", 75.0)

# MACD
MACD_FAST   = _ei("MACD_FAST", 12)
MACD_SLOW   = _ei("MACD_SLOW", 26)
MACD_SIG    = _ei("MACD_SIGNAL_PERIOD", 9)

# MTF
MTF_TIMEFRAME = os.getenv("MTF_TIMEFRAME", "1h")

# Dynamic sizing
DYN_SIZE_MIN    = _ef("DYNAMIC_SIZE_MIN",      5.0)
DYN_SIZE_MAX    = _ef("DYNAMIC_SIZE_MAX",      15.0)
DYN_AI_LOW      = _ef("DYNAMIC_SIZE_AI_LOW",   0.55)
DYN_AI_HIGH     = _ef("DYNAMIC_SIZE_AI_HIGH",  0.80)

# SL Cooldown
SL_COOLDOWN_N    = _ei("SL_COOLDOWN_AFTER_N",       2)
SL_PAUSE_SECONDS = _ei("SL_COOLDOWN_PAUSE_SECONDS", 1800)

# Breakeven
BREAKEVEN_TRIGGER = _ef("BREAKEVEN_TRIGGER_PCT", 0.5)

# OHLCV
TIMEFRAME    = os.getenv("BOT_TIMEFRAME", "15m")
CANDLE_LIMIT = _ei("BOT_CANDLE_LIMIT", 500)
ATR_PERIOD   = 14

# Walk-Forward
WF_TRAIN_PCT = 0.70   # 70% train, 30% test
WF_FOLDS     = 5

# Monte Carlo
MC_RUNS    = 1000
MC_CAPITAL = 1000.0   # starting capital USDT

# Parameter Optimization grid
PARAM_GRID = {
    "tp_pct":  [1.5, 2.0, 2.5, 3.0, 3.5, 4.0],
    "sl_pct":  [0.5, 0.75, 1.0, 1.25, 1.5],
    "rsi_min": [30, 35, 40, 45],
    "rsi_max": [65, 70, 75, 80],
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# OHLCV CACHE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_ohlcv_cache: Dict[str, list] = {}
_mtf_cache:   Dict[str, list] = {}


def fetch_ohlcv(symbol: str, timeframe: str = TIMEFRAME, limit: int = CANDLE_LIMIT) -> list:
    key = f"{symbol}_{timeframe}"
    cache = _ohlcv_cache if timeframe == TIMEFRAME else _mtf_cache
    if key in cache:
        return cache[key]
    try:
        import ccxt
        ex = ccxt.binance({"enableRateLimit": True})
        data = ex.fetch_ohlcv(symbol, timeframe=timeframe, limit=limit)
        cache[key] = data or []
        log.info(f"OHLCV | {symbol} {timeframe} candles={len(cache[key])}")
    except Exception as e:
        log.warning(f"OHLCV_FAIL | {symbol} {timeframe} err={e}")
        cache[key] = []
    return cache[key]


def ohlcv_to_df(ohlcv: list) -> pd.DataFrame:
    if not ohlcv:
        return pd.DataFrame()
    df = pd.DataFrame(ohlcv, columns=["ts", "open", "high", "low", "close", "volume"])
    df["ts"] = pd.to_datetime(df["ts"], unit="ms", utc=True)
    df = df.set_index("ts").sort_index()
    for c in ["open", "high", "low", "close", "volume"]:
        df[c] = df[c].astype(float)
    return df


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# INDICATOR LIBRARY (vectorbt-compatible — pandas Series output)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_atr(df: pd.DataFrame, n: int = ATR_PERIOD) -> pd.Series:
    """True ATR % of close price."""
    h, l, c = df["high"], df["low"], df["close"]
    pc  = c.shift(1)
    tr  = pd.concat([h - l, (h - pc).abs(), (l - pc).abs()], axis=1).max(axis=1)
    atr = tr.rolling(n).mean()
    return (atr / c * 100.0).rename("atr_pct")


def calc_rsi(close: pd.Series, n: int = RSI_PERIOD) -> pd.Series:
    """Wilder RSI."""
    delta = close.diff()
    gain  = delta.clip(lower=0)
    loss  = (-delta).clip(lower=0)
    avg_g = gain.ewm(com=n - 1, min_periods=n).mean()
    avg_l = loss.ewm(com=n - 1, min_periods=n).mean()
    rs    = avg_g / avg_l.replace(0, np.nan)
    return (100.0 - 100.0 / (1.0 + rs)).rename("rsi")


def calc_macd(close: pd.Series, fast: int = MACD_FAST,
              slow: int = MACD_SLOW, signal: int = MACD_SIG
              ) -> Tuple[pd.Series, pd.Series, pd.Series]:
    """MACD line, signal line, histogram."""
    ema_f = close.ewm(span=fast, adjust=False).mean()
    ema_s = close.ewm(span=slow, adjust=False).mean()
    macd  = (ema_f - ema_s).rename("macd")
    sig   = macd.ewm(span=signal, adjust=False).mean().rename("macd_signal")
    hist  = (macd - sig).rename("macd_hist")
    return macd, sig, hist


def calc_ema(close: pd.Series, n: int) -> pd.Series:
    return close.ewm(span=n, adjust=False).mean().rename(f"ema{n}")


def calc_sma(close: pd.Series, n: int) -> pd.Series:
    return close.rolling(n).mean().rename(f"sma{n}")


def calc_trend_strength(df: pd.DataFrame) -> pd.Series:
    """0..1 trend score — ბოტის _trend_strength-ის ემთხვევა."""
    c   = df["close"]
    s5  = calc_sma(c, 5)
    s10 = calc_sma(c, 10)
    slope = ((s5 / s10) - 1.0).fillna(0)
    mom1  = c.pct_change(1).fillna(0)
    ups3  = (c > c.shift(1)).rolling(3).sum().fillna(0) / 3.0

    base = (
        0.35 * (c > c.shift(1)).astype(float) +
        0.25 * (mom1  / 0.003).clip(0, 1) +
        0.20 * (slope / 0.003).clip(0, 1) +
        0.20 * ups3
    )
    return base.clip(0, 1).rename("trend")


def calc_volume_ratio(vol: pd.Series, n: int = 20) -> pd.Series:
    """Current volume / average volume."""
    return (vol / vol.rolling(n).mean()).fillna(1.0).rename("vol_ratio")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# REGIME DETECTION (vectorized)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_regime(trend: pd.Series, atr_pct: pd.Series) -> pd.Series:
    """Vectorized regime: BULL/UNCERTAIN/SIDEWAYS/BEAR/VOLATILE."""
    regime = pd.Series("UNCERTAIN", index=trend.index, dtype=str)
    regime[atr_pct >= VOLATILE_ATR_MIN]                                    = "VOLATILE"
    regime[(atr_pct <= SIDEWAYS_ATR_MAX) & (trend < BULL_TREND_MIN)]       = "SIDEWAYS"
    regime[(trend >= BULL_TREND_MIN) & (atr_pct < VOLATILE_ATR_MIN)]       = "BULL"
    regime[(trend <= BEAR_TREND_MAX) & (atr_pct < VOLATILE_ATR_MIN)]       = "BEAR"
    return regime.rename("regime")


def get_atr_tp_sl(regime: str, atr_pct: float) -> Tuple[float, float]:
    """ATR-based TP/SL — regime_engine-ს ემთხვევა."""
    mults = {
        "BULL":      (B_ATR_TP_BULL,      B_ATR_SL_BULL),
        "UNCERTAIN": (B_ATR_TP_UNCERTAIN, B_ATR_SL_UNCERTAIN),
    }
    if regime not in mults:
        return 0.0, 0.0
    tm, sm = mults[regime]
    tp = max(B_MIN_TP, min(B_MAX_TP, atr_pct * tm))
    sl = max(B_MIN_SL, min(B_MAX_SL, atr_pct * sm))
    return round(tp, 3), round(sl, 3)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SIGNAL GENERATION (vectorized — ყველა სტრატეგიისთვის)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_features(df: pd.DataFrame) -> pd.DataFrame:
    """ყველა indicator გამოითვლება ერთ DataFrame-ზე."""
    out = df.copy()
    out["atr_pct"]   = calc_atr(df)
    out["rsi"]       = calc_rsi(df["close"])
    out["trend"]     = calc_trend_strength(df)
    out["vol_ratio"] = calc_volume_ratio(df["volume"])
    out["regime"]    = calc_regime(out["trend"], out["atr_pct"])
    macd, sig, hist  = calc_macd(df["close"])
    out["macd"]      = macd
    out["macd_sig"]  = sig
    out["macd_hist"] = hist
    out["ema20"]     = calc_ema(df["close"], 20)
    out["ema50"]     = calc_ema(df["close"], 50)
    return out.dropna(subset=["atr_pct", "rsi", "trend"])


def signals_strategy_a(feat: pd.DataFrame) -> pd.Series:
    """Strategy A — ყველა candle BUY (baseline)."""
    return pd.Series(True, index=feat.index).rename("entry")


def signals_strategy_b(feat: pd.DataFrame) -> pd.Series:
    """Strategy B — Regime BULL/UNCERTAIN-ზე BUY."""
    return feat["regime"].isin(["BULL", "UNCERTAIN"]).rename("entry")


def signals_strategy_c(feat: pd.DataFrame,
                        rsi_min: float = RSI_MIN,
                        rsi_max: float = RSI_MAX) -> pd.Series:
    """Strategy C — RSI zone + MACD bullish crossover."""
    rsi_ok   = feat["rsi"].between(rsi_min, rsi_max)
    macd_ok  = feat["macd_hist"] > 0
    trend_ok = feat["trend"] >= 0.35
    return (rsi_ok & macd_ok & trend_ok).rename("entry")


def signals_strategy_d(feat: pd.DataFrame,
                        rsi_min: float = RSI_MIN,
                        rsi_max: float = RSI_MAX) -> pd.Series:
    """Strategy D — RSI + MACD + MTF (EMA20 > EMA50) + vol filter."""
    rsi_ok   = feat["rsi"].between(rsi_min, rsi_max)
    macd_ok  = feat["macd_hist"] > 0
    trend_ok = feat["trend"] >= BULL_TREND_MIN
    mtf_ok   = feat["ema20"] > feat["ema50"]
    vol_ok   = feat["vol_ratio"] >= 0.8
    regime_ok = feat["regime"].isin(["BULL", "UNCERTAIN"])
    return (rsi_ok & macd_ok & trend_ok & mtf_ok & vol_ok & regime_ok).rename("entry")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TRADE SIMULATOR (event-driven, one position at a time)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def simulate_trades(
    feat: pd.DataFrame,
    entries: pd.Series,
    tp_pct_fixed: Optional[float]  = None,
    sl_pct_fixed: Optional[float]  = None,
    use_dynamic_sizing: bool        = False,
    quote_size: float               = 7.0,
    use_breakeven: bool             = False,
    cooldown_n: int                 = SL_COOLDOWN_N,
    cooldown_pause: int             = SL_PAUSE_SECONDS,
    strategy_name: str              = "?",
) -> pd.DataFrame:
    """
    Event-driven trade simulator.
    ყველა სტრატეგიისთვის გამოიყენება — სხვადასხვა entry signals.

    Returns DataFrame: trade-ების სია PnL-ით.
    """
    trades: List[Dict[str, Any]] = []
    in_trade    = False
    entry_price = 0.0
    entry_ts    = None
    entry_atr   = 0.0
    entry_regime = ""
    entry_tp    = 0.0
    entry_sl    = 0.0
    entry_quote = quote_size
    peak_price  = 0.0

    # SL cooldown state
    consecutive_sl = 0
    sl_pause_until: Optional[pd.Timestamp] = None

    closes = feat["close"]
    idx    = feat.index

    for i, ts in enumerate(idx):
        row   = feat.iloc[i]
        price = float(closes.iloc[i])

        # ─── IN TRADE: check TP / SL / Breakeven ─────────────────
        if in_trade:
            # peak update (trailing / breakeven)
            if price > peak_price:
                peak_price = price

            # Breakeven — SL move to entry
            if use_breakeven:
                be_trigger = entry_price * (1.0 + BREAKEVEN_TRIGGER / 100.0)
                if peak_price >= be_trigger:
                    # SL გადაიწია entry-ზე
                    entry_sl_price = entry_price
                    sl_pct_eff     = 0.0  # SL at breakeven → PnL=0 (minus fees)
                else:
                    entry_sl_price = entry_price * (1.0 - entry_sl / 100.0)
                    sl_pct_eff     = entry_sl
            else:
                entry_sl_price = entry_price * (1.0 - entry_sl / 100.0)
                sl_pct_eff     = entry_sl

            tp_price = entry_price * (1.0 + entry_tp / 100.0)
            hit_tp   = price >= tp_price
            hit_sl   = price <= entry_sl_price

            if hit_tp or hit_sl:
                outcome   = "TP" if hit_tp else "SL"
                exit_pct  = entry_tp if hit_tp else sl_pct_eff
                gross     = entry_quote * (exit_pct / 100.0)
                fees      = entry_quote * COST
                pnl       = round((gross - fees) if hit_tp else -(gross + fees), 4)
                hold_bars = i - idx.get_loc(entry_ts)

                trades.append({
                    "strategy":     strategy_name,
                    "symbol":       row.get("symbol", ""),
                    "entry_ts":     entry_ts,
                    "exit_ts":      ts,
                    "entry_price":  round(entry_price, 6),
                    "exit_price":   round(price, 6),
                    "quote":        entry_quote,
                    "tp_pct":       entry_tp,
                    "sl_pct":       entry_sl,
                    "outcome":      outcome,
                    "pnl":          pnl,
                    "hold_bars":    hold_bars,
                    "regime":       entry_regime,
                    "atr_pct":      round(entry_atr, 4),
                    "rsi_entry":    round(float(feat.iloc[idx.get_loc(entry_ts)]["rsi"]) if "rsi" in feat.columns else 50.0, 1),
                    "macd_hist":    round(float(feat.iloc[idx.get_loc(entry_ts)]["macd_hist"]) if "macd_hist" in feat.columns else 0.0, 6),
                })

                # SL cooldown tracking
                if outcome == "SL":
                    consecutive_sl += 1
                    if consecutive_sl >= cooldown_n:
                        sl_pause_until = ts + pd.Timedelta(seconds=cooldown_pause)
                else:
                    consecutive_sl = 0
                    sl_pause_until = None

                in_trade = False
            continue

        # ─── NOT IN TRADE: check entry ────────────────────────────
        if not entries.iloc[i]:
            continue

        # SL Cooldown check
        if sl_pause_until is not None and ts < sl_pause_until:
            continue

        # Regime SKIP check
        regime = str(row.get("regime", "UNCERTAIN"))
        if regime in ("BEAR", "VOLATILE", "SIDEWAYS"):
            continue

        # TP/SL selection
        atr = float(row.get("atr_pct", 0.40) or 0.40)
        if tp_pct_fixed is not None:
            tp = tp_pct_fixed
            sl = sl_pct_fixed or A_SL_PCT
        else:
            tp, sl = get_atr_tp_sl(regime, atr)
            if tp == 0.0:
                continue

        # Minimum net edge check
        net = tp - (FEE_RT + SLIP)
        if net < B_MIN_TP:
            continue

        # Dynamic sizing
        if use_dynamic_sizing:
            ai_proxy = float(row.get("trend", 0.5))
            if ai_proxy <= DYN_AI_LOW:
                q = DYN_SIZE_MIN
            elif ai_proxy >= DYN_AI_HIGH:
                q = DYN_SIZE_MAX
            else:
                t_  = (ai_proxy - DYN_AI_LOW) / (DYN_AI_HIGH - DYN_AI_LOW)
                q   = DYN_SIZE_MIN + t_ * (DYN_SIZE_MAX - DYN_SIZE_MIN)
            q = min(round(q, 2), DYN_SIZE_MAX)
        else:
            q = quote_size

        # Open trade
        in_trade     = True
        entry_price  = price
        entry_ts     = ts
        entry_atr    = atr
        entry_regime = regime
        entry_tp     = tp
        entry_sl     = sl
        entry_quote  = q
        peak_price   = price

    return pd.DataFrame(trades)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# METRICS CALCULATOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_metrics(trades: pd.DataFrame, label: str = "") -> Dict[str, Any]:
    if trades.empty:
        return {k: 0 for k in [
            "strategy", "trades", "wins", "losses", "winrate",
            "total_pnl", "avg_win", "avg_loss", "profit_factor",
            "max_drawdown", "sharpe", "calmar", "expectancy",
            "avg_hold_bars", "roi_pct"
        ]}

    wins   = trades[trades["outcome"] == "TP"]
    losses = trades[trades["outcome"] == "SL"]
    n      = len(trades)
    wr     = len(wins) / n * 100 if n else 0.0

    avg_w  = float(wins["pnl"].mean())   if len(wins)   else 0.0
    avg_l  = float(losses["pnl"].abs().mean()) if len(losses) else 0.0
    pf     = avg_w / avg_l if avg_l > 0 else (float("inf") if avg_w > 0 else 0.0)

    total_pnl = float(trades["pnl"].sum())

    # Max drawdown
    cumul  = trades["pnl"].cumsum()
    peak   = cumul.cummax()
    dd     = float((cumul - peak).min())

    # Sharpe — daily PnL
    trades2 = trades.copy()
    trades2["date"] = pd.to_datetime(trades2["entry_ts"]).dt.date
    daily  = trades2.groupby("date")["pnl"].sum()
    if len(daily) >= 2:
        dr = pd.date_range(daily.index.min(), daily.index.max(), freq="D")
        daily = daily.reindex([d.date() for d in dr], fill_value=0.0)
    std    = float(daily.std()) if len(daily) > 1 else 0.0
    mean_d = float(daily.mean())
    sharpe = (mean_d / std * (252 ** 0.5)) if std > 0 else 0.0

    # Calmar
    calmar = (total_pnl / abs(dd)) if dd < 0 else float("inf")

    # Expectancy (per trade)
    expectancy = (wr / 100.0) * avg_w - (1 - wr / 100.0) * avg_l

    # Average hold bars
    avg_hold = float(trades["hold_bars"].mean()) if "hold_bars" in trades.columns else 0.0

    # ROI
    total_invested = float(trades["quote"].sum()) if "quote" in trades.columns else 1.0
    roi_pct = (total_pnl / total_invested * 100.0) if total_invested > 0 else 0.0

    return {
        "strategy":       label or str(trades["strategy"].iloc[0]) if "strategy" in trades.columns else "",
        "trades":         n,
        "wins":           len(wins),
        "losses":         len(losses),
        "winrate":        round(wr, 1),
        "total_pnl":      round(total_pnl, 4),
        "avg_win":        round(avg_w, 4),
        "avg_loss":       round(avg_l, 4),
        "profit_factor":  round(pf, 3),
        "max_drawdown":   round(dd, 4),
        "sharpe":         round(sharpe, 2),
        "calmar":         round(calmar, 3),
        "expectancy":     round(expectancy, 4),
        "avg_hold_bars":  round(avg_hold, 1),
        "roi_pct":        round(roi_pct, 3),
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MONTE CARLO SIMULATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def monte_carlo(trades: pd.DataFrame, runs: int = MC_RUNS,
                capital: float = MC_CAPITAL) -> Dict[str, Any]:
    """
    Trade PnL-ებს randomly reorder-ავს N ჯერ.
    რისკი, probability of ruin, expected range.
    """
    if trades.empty or len(trades) < 5:
        return {}

    pnls = trades["pnl"].values
    results = []

    rng = np.random.default_rng(42)
    for _ in range(runs):
        shuffled = rng.permutation(pnls)
        equity   = capital + np.cumsum(shuffled)
        peak     = np.maximum.accumulate(equity)
        dd       = float(np.min(equity - peak))
        final    = float(equity[-1])
        results.append({"final": final, "max_dd": dd})

    finals  = np.array([r["final"] for r in results])
    max_dds = np.array([r["max_dd"] for r in results])

    ruin_pct = 5.0  # ruin = lose > 5% of capital
    ruin = float(np.mean(finals < capital * (1.0 - ruin_pct / 100.0)) * 100.0)

    return {
        "mc_runs":          runs,
        "mc_median_final":  round(float(np.median(finals)), 2),
        "mc_p5_final":      round(float(np.percentile(finals, 5)), 2),
        "mc_p95_final":     round(float(np.percentile(finals, 95)), 2),
        "mc_median_max_dd": round(float(np.median(max_dds)), 2),
        "mc_p95_max_dd":    round(float(np.percentile(max_dds, 95)), 2),
        "mc_ruin_pct":      round(ruin, 1),
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WALK-FORWARD TESTING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def walk_forward_test(feat: pd.DataFrame, strategy_fn,
                      n_folds: int = WF_FOLDS,
                      train_pct: float = WF_TRAIN_PCT,
                      label: str = "?") -> Dict[str, Any]:
    """
    Rolling walk-forward: feature DataFrame-ს N fold-ზე ჭრის.
    ყოველ fold-ში: train set-ზე ოპტიმიზაცია, test set-ზე validation.
    """
    n      = len(feat)
    fold_n = n // n_folds
    fold_results = []

    for fold in range(n_folds):
        start   = fold * fold_n
        end_idx = min(start + fold_n, n)
        chunk   = feat.iloc[start:end_idx]

        split   = int(len(chunk) * train_pct)
        train   = chunk.iloc[:split]
        test    = chunk.iloc[split:]

        if len(test) < 10:
            continue

        # Train: optimize TP/SL grid
        best_pnl = -999.0
        best_tp  = A_TP_PCT
        best_sl  = A_SL_PCT

        for tp in [2.0, 2.5, 3.0, 3.5]:
            for sl in [0.75, 1.0, 1.25]:
                entries = strategy_fn(train)
                t       = simulate_trades(
                    train, entries,
                    tp_pct_fixed=tp, sl_pct_fixed=sl,
                    strategy_name=label,
                )
                if not t.empty and float(t["pnl"].sum()) > best_pnl:
                    best_pnl = float(t["pnl"].sum())
                    best_tp  = tp
                    best_sl  = sl

        # Test: validate with best params
        entries_test = strategy_fn(test)
        trades_test  = simulate_trades(
            test, entries_test,
            tp_pct_fixed=best_tp, sl_pct_fixed=best_sl,
            strategy_name=label,
        )
        m = calc_metrics(trades_test, label=f"{label}_fold{fold+1}")
        m["fold"]     = fold + 1
        m["best_tp"]  = best_tp
        m["best_sl"]  = best_sl
        fold_results.append(m)

    if not fold_results:
        return {}

    avg_wr  = float(np.mean([f["winrate"]      for f in fold_results]))
    avg_pnl = float(np.mean([f["total_pnl"]    for f in fold_results]))
    avg_pf  = float(np.mean([f["profit_factor"] for f in fold_results]))
    avg_sh  = float(np.mean([f["sharpe"]        for f in fold_results]))
    consistency = float(np.mean([1 if f["total_pnl"] > 0 else 0 for f in fold_results]) * 100)

    return {
        "wf_folds":       n_folds,
        "wf_avg_winrate": round(avg_wr, 1),
        "wf_avg_pnl":     round(avg_pnl, 4),
        "wf_avg_pf":      round(avg_pf, 3),
        "wf_avg_sharpe":  round(avg_sh, 2),
        "wf_consistency": round(consistency, 1),
        "wf_fold_details": fold_results,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PARAMETER OPTIMIZATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def optimize_params(feat: pd.DataFrame,
                    strategy: str = "C") -> pd.DataFrame:
    """
    Grid search on TP/SL/RSI params.
    Returns sorted DataFrame of all combinations.
    """
    log.info(f"Optimization: strategy={strategy} grid combinations={len(PARAM_GRID['tp_pct'])*len(PARAM_GRID['sl_pct'])*len(PARAM_GRID['rsi_min'])*len(PARAM_GRID['rsi_max'])}")

    results = []
    count   = 0

    for tp, sl, rmin, rmax in itertools.product(
        PARAM_GRID["tp_pct"],
        PARAM_GRID["sl_pct"],
        PARAM_GRID["rsi_min"],
        PARAM_GRID["rsi_max"],
    ):
        if rmin >= rmax:
            continue
        if strategy == "C":
            entries = signals_strategy_c(feat, rsi_min=rmin, rsi_max=rmax)
        elif strategy == "D":
            entries = signals_strategy_d(feat, rsi_min=rmin, rsi_max=rmax)
        else:
            entries = signals_strategy_b(feat)

        trades = simulate_trades(
            feat, entries,
            tp_pct_fixed=tp, sl_pct_fixed=sl,
            strategy_name=f"{strategy}_opt",
        )
        if trades.empty:
            continue

        m = calc_metrics(trades)
        results.append({
            "tp_pct":        tp,
            "sl_pct":        sl,
            "rsi_min":       rmin,
            "rsi_max":       rmax,
            "trades":        m["trades"],
            "winrate":       m["winrate"],
            "total_pnl":     m["total_pnl"],
            "profit_factor": m["profit_factor"],
            "sharpe":        m["sharpe"],
            "max_drawdown":  m["max_drawdown"],
        })
        count += 1

    log.info(f"Optimization complete | {count} param sets tested")
    df_opt = pd.DataFrame(results)
    if not df_opt.empty:
        df_opt = df_opt.sort_values("total_pnl", ascending=False).reset_index(drop=True)
    return df_opt


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# HTML REPORT GENERATOR
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_html_report(
    metrics_list: List[Dict],
    all_trades:   Dict[str, pd.DataFrame],
    mc_results:   Dict[str, Dict],
    wf_results:   Dict[str, Dict],
    output_path:  str = "backtest_v3_report.html",
) -> None:
    """Full HTML dashboard with equity curves, regime breakdown, metrics table."""

    def _color(val: float, good_positive: bool = True) -> str:
        if val > 0:
            return "#27ae60" if good_positive else "#e74c3c"
        elif val < 0:
            return "#e74c3c" if good_positive else "#27ae60"
        return "#888"

    def _bar(pct: float, max_w: int = 200) -> str:
        w = max(0, min(max_w, int(abs(pct) / 100.0 * max_w)))
        c = "#27ae60" if pct >= 0 else "#e74c3c"
        return f'<div style="background:{c};width:{w}px;height:12px;display:inline-block;border-radius:2px"></div>'

    # Build equity curve data
    equity_js = {}
    for name, trades in all_trades.items():
        if trades.empty:
            continue
        cumul = trades["pnl"].cumsum().tolist()
        labels = [str(ts)[:16] for ts in trades["entry_ts"].tolist()]
        equity_js[name] = {"labels": labels, "values": cumul}

    # Regime breakdown
    regime_rows = ""
    for name, trades in all_trades.items():
        if trades.empty or "regime" not in trades.columns:
            continue
        for regime, grp in trades.groupby("regime"):
            wins = (grp["outcome"] == "TP").sum()
            wr   = wins / len(grp) * 100 if len(grp) else 0
            pnl  = grp["pnl"].sum()
            regime_rows += f"""
            <tr>
              <td>{name}</td><td>{regime}</td>
              <td>{len(grp)}</td>
              <td style="color:{_color(wr-50)}">{wr:.1f}%</td>
              <td style="color:{_color(pnl)}">{pnl:+.4f}</td>
            </tr>"""

    # Metrics table rows
    metric_keys = [
        ("Trades",         "trades",        False),
        ("Wins",           "wins",          False),
        ("Losses",         "losses",        False),
        ("Winrate %",      "winrate",       True),
        ("Total PnL",      "total_pnl",     True),
        ("Avg Win",        "avg_win",       True),
        ("Avg Loss",       "avg_loss",      False),
        ("Profit Factor",  "profit_factor", True),
        ("Max Drawdown",   "max_drawdown",  False),
        ("Sharpe Ratio",   "sharpe",        True),
        ("Calmar Ratio",   "calmar",        True),
        ("Expectancy",     "expectancy",    True),
        ("ROI %",          "roi_pct",       True),
        ("Avg Hold Bars",  "avg_hold_bars", False),
    ]

    header_cells = "".join(f"<th>{m['strategy']}</th>" for m in metrics_list)
    metric_rows  = ""
    for label, key, gp in metric_keys:
        cells = f"<td><b>{label}</b></td>"
        for m in metrics_list:
            val = m.get(key, "-")
            if isinstance(val, float):
                color = _color(val, gp) if key not in ("avg_loss", "avg_hold_bars", "losses", "trades", "wins") else "#333"
                cells += f'<td style="color:{color};font-weight:500">{val}</td>'
            else:
                cells += f"<td>{val}</td>"
        metric_rows += f"<tr>{cells}</tr>"

    # Monte Carlo section
    mc_html = ""
    for name, mc in mc_results.items():
        if not mc:
            continue
        mc_html += f"""
        <div style="background:#f8f9fa;border-radius:8px;padding:16px;margin:8px;display:inline-block;min-width:220px">
          <h4 style="margin:0 0 8px">{name}</h4>
          <div>Runs: <b>{mc.get('mc_runs',0)}</b></div>
          <div>Median final: <b style="color:{_color(mc.get('mc_median_final',0)-MC_CAPITAL)}">${mc.get('mc_median_final',0):.2f}</b></div>
          <div>P5 / P95: <b>${mc.get('mc_p5_final',0):.2f}</b> / <b>${mc.get('mc_p95_final',0):.2f}</b></div>
          <div>Median max DD: <b style="color:#e74c3c">{mc.get('mc_median_max_dd',0):.2f}</b></div>
          <div>P95 max DD: <b style="color:#e74c3c">{mc.get('mc_p95_max_dd',0):.2f}</b></div>
          <div>Ruin probability: <b style="color:#e74c3c">{mc.get('mc_ruin_pct',0):.1f}%</b></div>
        </div>"""

    # Walk-Forward section
    wf_html = ""
    for name, wf in wf_results.items():
        if not wf:
            continue
        cons_color = _color(wf.get("wf_consistency", 0) - 50)
        wf_html += f"""
        <div style="background:#f8f9fa;border-radius:8px;padding:16px;margin:8px;display:inline-block;min-width:220px">
          <h4 style="margin:0 0 8px">{name}</h4>
          <div>Folds: <b>{wf.get('wf_folds',0)}</b></div>
          <div>Avg Winrate: <b>{wf.get('wf_avg_winrate',0):.1f}%</b></div>
          <div>Avg PnL/fold: <b style="color:{_color(wf.get('wf_avg_pnl',0))}">{wf.get('wf_avg_pnl',0):+.4f}</b></div>
          <div>Avg Sharpe: <b>{wf.get('wf_avg_sharpe',0):.2f}</b></div>
          <div>Consistency: <b style="color:{cons_color}">{wf.get('wf_consistency',0):.1f}%</b></div>
        </div>"""

    # Best strategy
    best = max(metrics_list, key=lambda m: m.get("total_pnl", -999)) if metrics_list else {}
    best_name = best.get("strategy", "?")
    best_pnl  = best.get("total_pnl", 0)

    html = f"""<!DOCTYPE html>
<html lang="ka">
<head>
<meta charset="UTF-8">
<title>Genius Bot Backtest v3</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
          background: #0f1117; color: #e0e0e0; padding: 24px; }}
  h1 {{ font-size: 28px; font-weight: 700; margin-bottom: 4px; }}
  h2 {{ font-size: 18px; font-weight: 600; margin: 24px 0 12px;
        border-bottom: 1px solid #333; padding-bottom: 6px; }}
  .subtitle {{ color: #888; font-size: 14px; margin-bottom: 24px; }}
  .card {{ background: #1a1d27; border-radius: 12px; padding: 20px; margin-bottom: 20px; }}
  .badge {{ display: inline-block; padding: 4px 12px; border-radius: 20px;
             font-size: 13px; font-weight: 600; }}
  .badge-green {{ background: #1a3a2a; color: #27ae60; }}
  .badge-red   {{ background: #3a1a1a; color: #e74c3c; }}
  .badge-blue  {{ background: #1a2a3a; color: #3498db; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  th {{ background: #252836; padding: 10px 12px; text-align: left;
        font-weight: 600; color: #aaa; font-size: 13px; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #1e2130; }}
  tr:hover td {{ background: #1e2130; }}
  .chart-wrap {{ position: relative; height: 280px; }}
  .flex {{ display: flex; flex-wrap: wrap; gap: 12px; }}
  .metric-card {{ background: #252836; border-radius: 8px; padding: 14px 18px;
                  min-width: 160px; }}
  .metric-label {{ font-size: 12px; color: #888; margin-bottom: 4px; }}
  .metric-value {{ font-size: 22px; font-weight: 700; }}
</style>
</head>
<body>

<h1>🤖 Genius Bot — Backtest v3</h1>
<div class="subtitle">
  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} |
  Fee={FEE_RT}% Slip={SLIP}% |
  Starting capital: ${MC_CAPITAL:.0f} USDT
</div>

<!-- VERDICT -->
<div class="card">
  <h2>🏆 Best Strategy</h2>
  <div class="flex" style="align-items:center;gap:20px">
    <div class="metric-card">
      <div class="metric-label">Winner</div>
      <div class="metric-value" style="color:#f39c12">{best_name}</div>
    </div>
    <div class="metric-card">
      <div class="metric-label">Total PnL</div>
      <div class="metric-value" style="color:{'#27ae60' if best_pnl>0 else '#e74c3c'}">{best_pnl:+.4f} USDT</div>
    </div>
    {"".join(
        '<div class="metric-card"><div class="metric-label">' + m["strategy"] +
        '</div><div class="metric-value" style="font-size:16px;color:' +
        _color(m.get("total_pnl", 0)) + '">' +
        f'{m.get("total_pnl", 0):+.4f}' +
        '</div></div>'
        for m in metrics_list
    )}
  </div>
</div>

<!-- METRICS TABLE -->
<div class="card">
  <h2>📊 Strategy Comparison</h2>
  <table>
    <thead><tr><th>Metric</th>{header_cells}</tr></thead>
    <tbody>{metric_rows}</tbody>
  </table>
</div>

<!-- EQUITY CURVES -->
<div class="card">
  <h2>📈 Equity Curves</h2>
  <div class="chart-wrap">
    <canvas id="equityChart"></canvas>
  </div>
</div>

<!-- REGIME BREAKDOWN -->
<div class="card">
  <h2>🎯 Regime Breakdown</h2>
  <table>
    <thead><tr><th>Strategy</th><th>Regime</th><th>Trades</th><th>Winrate</th><th>PnL</th></tr></thead>
    <tbody>{regime_rows}</tbody>
  </table>
</div>

<!-- MONTE CARLO -->
<div class="card">
  <h2>🎲 Monte Carlo ({MC_RUNS} runs, capital=${MC_CAPITAL:.0f})</h2>
  <div class="flex">{mc_html}</div>
</div>

<!-- WALK-FORWARD -->
<div class="card">
  <h2>🔄 Walk-Forward Testing ({WF_FOLDS} folds)</h2>
  <div class="flex">{wf_html}</div>
</div>

<script>
const eq = {json.dumps(equity_js)};
const colors = ['#3498db','#27ae60','#f39c12','#e74c3c','#9b59b6'];
const datasets = Object.entries(eq).map(([name, d], i) => ({{
  label: name,
  data: d.values,
  borderColor: colors[i % colors.length],
  backgroundColor: 'transparent',
  borderWidth: 2,
  pointRadius: 0,
  tension: 0.3,
}}));
const allLabels = Object.values(eq)[0]?.labels || [];
new Chart(document.getElementById('equityChart'), {{
  type: 'line',
  data: {{ labels: allLabels, datasets }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ labels: {{ color: '#ccc' }} }} }},
    scales: {{
      x: {{ ticks: {{ color: '#666', maxTicksLimit: 12 }}, grid: {{ color: '#1e2130' }} }},
      y: {{ ticks: {{ color: '#aaa' }}, grid: {{ color: '#1e2130' }},
           title: {{ display: true, text: 'PnL USDT', color: '#888' }} }}
    }}
  }}
}});
</script>
</body></html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"HTML report saved: {output_path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL EXPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def save_excel_report(
    all_trades:   Dict[str, pd.DataFrame],
    metrics_list: List[Dict],
    opt_df:       Optional[pd.DataFrame],
    output_path:  str = "backtest_v3_results.xlsx",
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb      = Workbook()
        green_f = PatternFill("solid", start_color="C6EFCE")
        red_f   = PatternFill("solid", start_color="FFC7CE")
        hdr_f   = PatternFill("solid", start_color="1F4E79")
        hdr_font= Font(bold=True, color="FFFFFF")

        def _hdr(ws, headers):
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font  = hdr_font
                cell.fill  = hdr_f
                cell.alignment = Alignment(horizontal="center")

        # Summary sheet
        ws_sum = wb.active
        ws_sum.title = "Summary"
        sum_headers = ["Metric"] + [m["strategy"] for m in metrics_list]
        _hdr(ws_sum, sum_headers)
        metric_keys2 = [
            "trades","wins","losses","winrate","total_pnl",
            "avg_win","avg_loss","profit_factor","max_drawdown",
            "sharpe","calmar","expectancy","roi_pct","avg_hold_bars",
        ]
        for r, key in enumerate(metric_keys2, 2):
            ws_sum.cell(r, 1, key)
            for c, m in enumerate(metrics_list, 2):
                val = m.get(key, "")
                cell = ws_sum.cell(r, c, val)
                if isinstance(val, float) and key in ("total_pnl","sharpe","profit_factor"):
                    cell.fill = green_f if val > 0 else red_f

        # Per-strategy trade sheets
        trade_cols = ["entry_ts","exit_ts","entry_price","exit_price",
                      "quote","tp_pct","sl_pct","outcome","pnl",
                      "regime","atr_pct","rsi_entry","hold_bars"]
        for name, trades in all_trades.items():
            if trades.empty:
                continue
            ws = wb.create_sheet(name[:28])
            available = [c for c in trade_cols if c in trades.columns]
            _hdr(ws, available)
            for r, (_, row) in enumerate(trades.iterrows(), 2):
                for c, col in enumerate(available, 1):
                    val = row.get(col, "")
                    if isinstance(val, float) and math.isnan(val):
                        val = ""
                    ws.cell(r, c, val)
                pnl_val = row.get("pnl", 0) or 0
                pnl_col = available.index("pnl") + 1 if "pnl" in available else None
                if pnl_col:
                    ws.cell(r, pnl_col).fill = green_f if pnl_val > 0 else red_f

        # Optimization sheet
        if opt_df is not None and not opt_df.empty:
            ws_opt = wb.create_sheet("Optimization")
            _hdr(ws_opt, list(opt_df.columns))
            for r, (_, row) in enumerate(opt_df.iterrows(), 2):
                for c, col in enumerate(opt_df.columns, 1):
                    ws_opt.cell(r, c, row[col])
                if "total_pnl" in opt_df.columns:
                    pnl_c = list(opt_df.columns).index("total_pnl") + 1
                    ws_opt.cell(r, pnl_c).fill = green_f if row["total_pnl"] > 0 else red_f

        wb.save(output_path)
        log.info(f"Excel saved: {output_path}")
    except Exception as e:
        log.warning(f"Excel save failed: {e}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONSOLE REPORT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def print_console_report(metrics_list: List[Dict], mc_results: Dict,
                          wf_results: Dict) -> None:
    W = 72
    print("\n" + "═" * W)
    print("  🤖 GENIUS BOT — BACKTEST v3 RESULTS")
    print(f"  Fee={FEE_RT}%  Slip={SLIP}%  Cost/trade={(FEE_RT+SLIP):.3f}%")
    print("═" * W)

    col_w = (W - 24) // max(len(metrics_list), 1)
    header = f"  {'Metric':<22}" + "".join(f"{m['strategy']:>{col_w}}" for m in metrics_list)
    print(header)
    print("─" * W)

    rows = [
        ("Trades",         "trades"),
        ("Wins (TP)",      "wins"),
        ("Losses (SL)",    "losses"),
        ("Winrate %",      "winrate"),
        ("Total PnL USDT", "total_pnl"),
        ("Avg Win",        "avg_win"),
        ("Avg Loss",       "avg_loss"),
        ("Profit Factor",  "profit_factor"),
        ("Max Drawdown",   "max_drawdown"),
        ("Sharpe Ratio",   "sharpe"),
        ("Calmar Ratio",   "calmar"),
        ("Expectancy",     "expectancy"),
        ("ROI %",          "roi_pct"),
    ]
    for label, key in rows:
        line = f"  {label:<22}"
        for m in metrics_list:
            v = m.get(key, "-")
            line += f"{str(v):>{col_w}}"
        print(line)

    print("═" * W)
    best = max(metrics_list, key=lambda m: m.get("total_pnl", -999))
    print(f"\n  🏆 BEST STRATEGY : {best['strategy']}")
    print(f"     Total PnL     : {best.get('total_pnl',0):+.4f} USDT")
    print(f"     Winrate       : {best.get('winrate',0):.1f}%")
    print(f"     Sharpe        : {best.get('sharpe',0):.2f}")

    # WF summary
    for name, wf in wf_results.items():
        if wf:
            print(f"\n  Walk-Forward [{name}]: avg_wr={wf.get('wf_avg_winrate',0):.1f}% "
                  f"consistency={wf.get('wf_consistency',0):.1f}% "
                  f"avg_pnl={wf.get('wf_avg_pnl',0):+.4f}")

    # MC summary
    for name, mc in mc_results.items():
        if mc:
            print(f"  Monte Carlo  [{name}]: median=${mc.get('mc_median_final',0):.2f} "
                  f"ruin={mc.get('mc_ruin_pct',0):.1f}% "
                  f"p95_dd={mc.get('mc_p95_max_dd',0):.2f}")
    print()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN RUNNERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def run_ohlcv_mode(symbols: List[str]) -> None:
    """ყველა 4 სტრატეგია OHLCV data-ზე."""
    log.info(f"OHLCV mode | symbols={symbols} timeframe={TIMEFRAME}")

    all_feat: Dict[str, pd.DataFrame] = {}
    for sym in symbols:
        raw = fetch_ohlcv(sym, TIMEFRAME, CANDLE_LIMIT)
        df  = ohlcv_to_df(raw)
        if df.empty:
            log.warning(f"No OHLCV data for {sym}")
            continue
        feat = build_features(df)
        feat["symbol"] = sym
        all_feat[sym]  = feat
        log.info(f"Features built | {sym} rows={len(feat)}")

    if not all_feat:
        log.error("No feature data. Exiting.")
        return

    # Combine all symbols
    combined = pd.concat(all_feat.values(), ignore_index=False)
    combined = combined.sort_index()

    # Run 4 strategies
    strategies = {
        "A_Fixed":     (signals_strategy_a, {"tp_pct_fixed": A_TP_PCT, "sl_pct_fixed": A_SL_PCT}),
        "B_Regime":    (signals_strategy_b, {}),
        "C_RSI_MACD":  (signals_strategy_c, {}),
        "D_Full":      (signals_strategy_d, {"use_dynamic_sizing": True}),
    }

    all_trades:   Dict[str, pd.DataFrame] = {}
    metrics_list: List[Dict]              = []
    mc_results:   Dict[str, Dict]         = {}
    wf_results:   Dict[str, Dict]         = {}

    for name, (sig_fn, kwargs) in strategies.items():
        log.info(f"Simulating {name}...")
        entries = sig_fn(combined)
        trades  = simulate_trades(
            combined, entries, strategy_name=name, **kwargs
        )
        all_trades[name] = trades
        m = calc_metrics(trades, label=name)
        metrics_list.append(m)
        log.info(f"  {name}: trades={m['trades']} wr={m['winrate']}% pnl={m['total_pnl']:+.4f}")

        # Monte Carlo on best 2 strategies
        if name in ("C_RSI_MACD", "D_Full"):
            mc_results[name] = monte_carlo(trades)

        # Walk-Forward on main strategies
        if name in ("B_Regime", "D_Full"):
            wf_results[name] = walk_forward_test(combined, sig_fn, label=name)

    print_console_report(metrics_list, mc_results, wf_results)
    generate_html_report(metrics_list, all_trades, mc_results, wf_results)
    save_excel_report(all_trades, metrics_list, opt_df=None)


def run_history_mode(filepath: str) -> None:
    """Binance Order History Excel — ძველი backtest.py-ს ფუნქციონალი + metrics."""
    log.info(f"History mode | file={filepath}")

    try:
        df = pd.read_excel(filepath)
    except Exception as e:
        log.error(f"Cannot read file: {e}")
        return

    for col in ["Total", "AvgTrading Price", "Filled"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    real = df[df["Type"].isin(["BUY","SELL"]) & (df["Status"] == "Filled")].copy()
    real["Date(UTC)"] = pd.to_datetime(real["Date(UTC)"], errors="coerce")
    real = real.dropna(subset=["Date(UTC)", "AvgTrading Price"]).sort_values("Date(UTC)").reset_index(drop=True)

    symbols = real["Pair"].dropna().unique().tolist()
    log.info(f"Symbols detected: {symbols}")

    ohlcv_map = {sym: fetch_ohlcv(sym) for sym in symbols}

    pairs = []
    for sym in symbols:
        sym_df = real[real["Pair"] == sym].reset_index(drop=True)
        ohlcv  = ohlcv_map.get(sym, [])
        i = 0
        while i < len(sym_df):
            row = sym_df.iloc[i]
            if row["Type"] != "BUY":
                i += 1; continue
            sell_idx = None
            for j in range(i + 1, len(sym_df)):
                if sym_df.iloc[j]["Type"] == "SELL":
                    sell_idx = j; break
                if sym_df.iloc[j]["Type"] == "BUY":
                    break
            if sell_idx is None:
                i += 1; continue
            nxt        = sym_df.iloc[sell_idx]
            buy_price  = float(row["AvgTrading Price"])
            sell_price = float(nxt["AvgTrading Price"])
            quote      = float(row["Total"])
            if buy_price <= 0 or quote <= 0:
                i = sell_idx + 1; continue
            actual_pct = (sell_price - buy_price) / buy_price * 100
            buy_ts_ms  = int(row["Date(UTC)"].timestamp() * 1000)
            if ohlcv:
                past    = [c for c in ohlcv if int(c[0]) <= buy_ts_ms]
                atr_pct = _atr_from_history(past)
                trend   = _trend_from_history(past)
            else:
                atr_pct = 0.40
                trend   = 0.40
            regime = "BULL" if trend >= BULL_TREND_MIN else "UNCERTAIN"
            tp, sl = get_atr_tp_sl(regime, atr_pct)
            outcome_real = "SL" if sell_price < buy_price else "TP"
            pairs.append({
                "entry_ts":    row["Date(UTC)"],
                "exit_ts":     nxt["Date(UTC)"],
                "entry_price": buy_price,
                "exit_price":  sell_price,
                "quote":       quote,
                "actual_pct":  round(actual_pct, 4),
                "outcome":     outcome_real,
                "tp_pct":      tp,
                "sl_pct":      sl,
                "pnl":         round(
                    quote * (abs(actual_pct) / 100.0) - quote * COST
                    if outcome_real == "TP"
                    else -(quote * (abs(actual_pct) / 100.0) + quote * COST),
                    4),
                "regime":      regime,
                "atr_pct":     round(atr_pct, 4),
                "hold_bars":   0,
                "strategy":    "History",
            })
            i = sell_idx + 1

    if not pairs:
        log.error("No trade pairs found.")
        return

    trades = pd.DataFrame(pairs)
    m = calc_metrics(trades, label="History")
    mc = monte_carlo(trades)
    print_console_report([m], {}, {})
    generate_html_report([m], {"History": trades}, {"History": mc}, {})
    save_excel_report({"History": trades}, [m], opt_df=None)


def _atr_from_history(ohlcv: list, n: int = ATR_PERIOD) -> float:
    if len(ohlcv) < n + 1:
        return 0.40
    trs = []
    for i in range(-n, 0):
        h = float(ohlcv[i][2]); l = float(ohlcv[i][3]); pc = float(ohlcv[i-1][4])
        trs.append(max(h-l, abs(h-pc), abs(l-pc)))
    atr = sum(trs) / n
    lc  = float(ohlcv[-1][4])
    return (atr / lc * 100.0) if lc > 0 else 0.40


def _trend_from_history(ohlcv: list) -> float:
    if len(ohlcv) < 10:
        return 0.40
    closes = [float(c[4]) for c in ohlcv]
    s5  = sum(closes[-5:]) / 5
    s10 = sum(closes[-10:]) / 10
    if s10 == 0:
        return 0.40
    slope = (s5/s10) - 1.0
    mom1  = (closes[-1]/closes[-2] - 1.0) if closes[-2] else 0.0
    ups3  = sum(1 for i in range(-3, 0) if closes[i] > closes[i-1])
    base  = 0.35*(1.0 if closes[-1]>closes[-2] else 0) + 0.25*max(0,min(1,mom1/0.003)) + 0.20*max(0,min(1,slope/0.003)) + 0.20*(ups3/3)
    return max(0.0, min(1.0, base))


def run_optimize_mode(symbols: List[str]) -> None:
    """Parameter optimization — best TP/SL/RSI combination."""
    log.info(f"Optimize mode | symbols={symbols}")
    combined = pd.concat(
        [build_features(ohlcv_to_df(fetch_ohlcv(s))).assign(symbol=s) for s in symbols],
        ignore_index=False
    ).sort_index()
    if combined.empty:
        log.error("No data"); return

    for strat in ["C", "D"]:
        log.info(f"Optimizing strategy {strat}...")
        opt = optimize_params(combined, strategy=strat)
        if not opt.empty:
            print(f"\nTop 10 params for Strategy {strat}:")
            print(opt.head(10).to_string(index=False))
            opt.to_csv(f"optimization_{strat}.csv", index=False)
            log.info(f"Saved: optimization_{strat}.csv")


def run_walkforward_mode(symbols: List[str]) -> None:
    """Walk-forward testing."""
    log.info(f"Walk-forward mode | symbols={symbols}")
    combined = pd.concat(
        [build_features(ohlcv_to_df(fetch_ohlcv(s))).assign(symbol=s) for s in symbols],
        ignore_index=False
    ).sort_index()
    if combined.empty:
        log.error("No data"); return

    wf_results = {}
    for name, fn in [("B_Regime", signals_strategy_b), ("D_Full", signals_strategy_d)]:
        wf = walk_forward_test(combined, fn, label=name)
        wf_results[name] = wf
        if wf:
            print(f"\n{name} Walk-Forward:")
            for fd in wf.get("wf_fold_details", []):
                print(f"  Fold {fd['fold']}: wr={fd['winrate']}% pnl={fd['total_pnl']:+.4f} tp={fd['best_tp']} sl={fd['best_sl']}")
            print(f"  Summary: avg_wr={wf['wf_avg_winrate']:.1f}% consistency={wf['wf_consistency']:.1f}%")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ENTRY POINT
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def main() -> None:
    global TIMEFRAME, CANDLE_LIMIT

    parser = argparse.ArgumentParser(description="Genius Bot Backtest v3")
    parser.add_argument("--mode", choices=["ohlcv","history","optimize","walkforward"],
                        default="ohlcv")
    parser.add_argument("--symbols", nargs="+",
                        default=["BTC/USDT", "ETH/USDT", "BNB/USDT"])
    parser.add_argument("--file",   default="",
                        help="Binance Order History xlsx (history mode)")
    parser.add_argument("--timeframe", default=TIMEFRAME)
    parser.add_argument("--candles",   type=int, default=CANDLE_LIMIT)
    args = parser.parse_args()

    TIMEFRAME    = args.timeframe
    CANDLE_LIMIT = args.candles

    log.info(f"Genius Bot Backtest v3 | mode={args.mode} symbols={args.symbols}")

    if args.mode == "ohlcv":
        run_ohlcv_mode(args.symbols)
    elif args.mode == "history":
        if not args.file:
            log.error("--file required for history mode")
            sys.exit(1)
        run_history_mode(args.file)
    elif args.mode == "optimize":
        run_optimize_mode(args.symbols)
    elif args.mode == "walkforward":
        run_walkforward_mode(args.symbols)


if __name__ == "__main__":
    main()
